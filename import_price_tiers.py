#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse /tmp/eve/2.6.5-0123-work.xlsx sheet 熟豆豆单 A1:P62
and generate SQL to upsert products + insert tiers.

Assumptions:
- Bags are 454g = 1 lb (noted in sheet).
- Tier columns are like '2包', '14包', ... meaning min quantity in bags/lb.
- For each product, use the '折后价' row if present; else use the product row.
- If a tier cell is '-' or empty, skip that tier.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, List, Optional, Tuple

from openpyxl import load_workbook

SCHEMA = "p2rms15pepb5ciz"
SHEET = "熟豆豆单"
RANGE_MIN_ROW = 1
RANGE_MAX_ROW = 62
RANGE_MIN_COL = 1
RANGE_MAX_COL = 16


@dataclass
class Tier:
    min_lb: float
    max_lb: Optional[float]
    price_per_lb: float


def norm_name(s: str) -> str:
    s = (s or "").strip()
    # First line is usually the canonical name
    s = s.split("\n", 1)[0].strip()
    # Remove trailing punctuation/spaces
    s = re.sub(r"\s+", " ", s)
    return s


def parse_tier_header(vals: List[Any]) -> List[Tuple[int, int]]:
    """Return list of (colIndex0, qty_lb_int) for columns representing tiers.
    Input vals is the full row (A..P)."""
    out: List[Tuple[int, int]] = []
    for idx, v in enumerate(vals):
        if not isinstance(v, str):
            continue
        m = re.fullmatch(r"\s*(\d+)\s*包\s*", v)
        if not m:
            continue
        out.append((idx, int(m.group(1))))
    out.sort(key=lambda x: x[1])
    return out


def as_price(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        t = v.strip()
        if t in ("", "-", "—"):
            return None
        # sometimes string numbers
        try:
            return float(t)
        except Exception:
            return None
    return None


def find_sections(rows: List[List[Any]]) -> List[int]:
    """Return row indices (0-based) where a tier table header begins (the row containing 序号/名称 and 包 columns)."""
    starts = []
    for i, r in enumerate(rows):
        if len(r) < 6:
            continue
        # Header pattern: col D='序号', col E='名称'
        if (isinstance(r[3], str) and r[3].strip() == "序号") and (isinstance(r[4], str) and r[4].strip() == "名称"):
            starts.append(i)
    return starts


def extract_products(rows: List[List[Any]]):
    """Yield dict(name -> tiers)"""
    sections = find_sections(rows)
    for si, start in enumerate(sections):
        end = sections[si + 1] if si + 1 < len(sections) else len(rows)
        header = rows[start]
        tier_cols = parse_tier_header(header)
        if not tier_cols:
            continue

        i = start + 1
        while i < end:
            r = rows[i]
            seq = r[3]
            name_cell = r[4]
            # product row: seq is number and name is str
            if isinstance(seq, (int, float)) and isinstance(name_cell, str) and name_cell.strip() not in ("", "折后价"):
                name = norm_name(name_cell)
                # determine price row: prefer next row if it is 折后价
                price_row = r
                if i + 1 < end and isinstance(rows[i + 1][4], str) and rows[i + 1][4].strip() == "折后价":
                    price_row = rows[i + 1]

                # collect (qty_lb, price)
                pts = []
                for col0, qty in tier_cols:
                    p = as_price(price_row[col0])
                    if p is None:
                        continue
                    pts.append((qty, p))

                # build tiers min/max from sorted qty
                pts.sort(key=lambda x: x[0])
                tiers: List[Tier] = []
                for j, (minq, price) in enumerate(pts):
                    maxq = None
                    if j + 1 < len(pts):
                        maxq = float(pts[j + 1][0] - 1)
                    tiers.append(Tier(min_lb=float(minq), max_lb=maxq, price_per_lb=float(price)))

                if tiers:
                    yield name, tiers

                # skip possible discount row
                if price_row is rows[i + 1] if i + 1 < end else False:
                    i += 2
                else:
                    i += 1
                continue

            i += 1


def sql_escape(s: str) -> str:
    return s.replace("'", "''")


def main():
    wb = load_workbook("2.6.5-0123-work.xlsx", data_only=True)
    sh = wb[SHEET]

    rows: List[List[Any]] = []
    for r in sh.iter_rows(min_row=RANGE_MIN_ROW, max_row=RANGE_MAX_ROW, min_col=RANGE_MIN_COL, max_col=RANGE_MAX_COL, values_only=True):
        rows.append(list(r))

    products = {}
    for name, tiers in extract_products(rows):
        products[name] = tiers

    print(f"-- products: {len(products)}")
    print("BEGIN;")
    print("SET LOCAL lock_timeout = '5s';")

    # Upsert products
    for name in sorted(products.keys()):
        esc = sql_escape(name)
        print(f"INSERT INTO {SCHEMA}.products(name, default_price, active) VALUES ('{esc}', 0, TRUE) ON CONFLICT (name) DO NOTHING;")

    # Insert tiers (idempotent-ish: delete and recreate for names in import)
    print(f"-- refresh tiers for imported products")
    print(f"DELETE FROM {SCHEMA}.product_price_tiers t USING {SCHEMA}.products p WHERE t.product_id=p.id AND p.name IN (")
    names = sorted(products.keys())
    for i, name in enumerate(names):
        esc = sql_escape(name)
        comma = "," if i + 1 < len(names) else ""
        print(f"  '{esc}'{comma}")
    print(");")

    for name in names:
        tiers = products[name]
        esc = sql_escape(name)
        print(f"-- {esc}")
        for t in tiers:
            maxv = "NULL" if t.max_lb is None else str(t.max_lb)
            print(
                f"INSERT INTO {SCHEMA}.product_price_tiers(product_id,min_qty_lb,max_qty_lb,price_per_lb,active) "
                f"SELECT id,{t.min_lb},{maxv},{t.price_per_lb},TRUE FROM {SCHEMA}.products WHERE name='{esc}';"
            )

    print("COMMIT;")


if __name__ == '__main__':
    main()
